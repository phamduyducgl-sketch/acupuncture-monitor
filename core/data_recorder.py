"""
Ghi lưu dữ liệu châm kim vào file Excel.
Mỗi sinh viên có 1 file riêng theo MSSV.
"""

import os
import shutil
import openpyxl
from typing import Optional


DATA_DIR     = os.path.join(os.path.dirname(__file__), "data")
TEMPLATE_PATH = os.path.join(DATA_DIR, "sample.xlsx")
COLUMNS      = ["", "VELOCITY (cm/s)", "ANGLE (deg)", "LENGTH (cm)"]


class DataRecorder:
    """
    Quản lý file Excel kết quả châm kim.

    Sử dụng:
        recorder = DataRecorder()
        recorder.create_session("SV001")
        recorder.record(velocity=2.5, angle=45, length=1.2)
        path = recorder.file_path   # → .../data/SV001.xlsx
    """

    def __init__(self, data_dir: str = DATA_DIR):
        self._data_dir = data_dir
        self._student_id: Optional[str] = None
        self._workbook = None
        self._worksheet = None
        os.makedirs(self._data_dir, exist_ok=True)
        self._ensure_template()

    # ── Public API ───────────────────────────────────────────────────────────

    def create_session(self, student_id: str):
        """
        Tạo file Excel mới cho sinh viên.
        Nếu file đã tồn tại, xoá sạch dữ liệu cũ.
        """
        self._student_id = student_id
        dest = self._path_for(student_id)
        shutil.copy(TEMPLATE_PATH, dest)

        self._workbook  = openpyxl.load_workbook(dest)
        self._worksheet = self._workbook.active

        # Xoá toàn bộ hàng cũ
        self._worksheet.delete_rows(1, self._worksheet.max_row + 1)
        # Tạo lại tiêu đề cột
        self._worksheet.append(COLUMNS)
        self._workbook.save(dest)

    def record(self, velocity: float, angle: int, length: float):
        """Ghi 1 lần đo vào file."""
        if not self._student_id:
            raise RuntimeError("Chưa tạo session — hãy gọi create_session() trước.")
        self._worksheet.append(["", velocity, angle, length])
        self._workbook.save(self._path_for(self._student_id))

    @property
    def file_path(self) -> Optional[str]:
        """Đường dẫn file Excel hiện tại."""
        if self._student_id:
            return self._path_for(self._student_id)
        return None

    @property
    def student_id(self) -> Optional[str]:
        return self._student_id

    # ── Internal ─────────────────────────────────────────────────────────────

    def _path_for(self, student_id: str) -> str:
        return os.path.join(self._data_dir, f"{student_id}.xlsx")

    def _ensure_template(self):
        if not os.path.exists(TEMPLATE_PATH):
            wb = openpyxl.Workbook()
            wb.save(TEMPLATE_PATH)
